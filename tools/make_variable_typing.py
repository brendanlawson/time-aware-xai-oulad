"""Generate docs/variable_typing.xlsx (STT 16, Duc).

Bilingual variable-typing catalogue derived from the canonical types in
src/features/preprocessing.py and the data dictionary. One row per variable with
its group, type (EN + VI), ordinal order, encoding, and origin.

Run with a Python that has openpyxl:  python tools/make_variable_typing.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "docs" / "variable_typing.xlsx"

# Variable | Group | Type(EN) | Kiểu(VI) | Order (if ordinal) | Encoding | Origin
ROWS = [
    (
        "id_student",
        "Identifier",
        "Nominal (ID)",
        "Định danh",
        "",
        "Group key (GroupKFold)",
        "Original",
    ),
    (
        "code_module",
        "Identifier",
        "Nominal (ID)",
        "Định danh",
        "",
        "Composite key",
        "Original",
    ),
    (
        "code_presentation",
        "Identifier",
        "Nominal (ID)",
        "Định danh",
        "",
        "Composite key",
        "Original",
    ),
    (
        "gender",
        "Demographic",
        "Binary",
        "Nhị phân",
        "",
        "Binary 0/1 (M=1,F=0)",
        "Original",
    ),
    ("region", "Demographic", "Nominal", "Danh định", "", "One-Hot", "Original"),
    (
        "highest_education",
        "Demographic",
        "Ordinal",
        "Thứ bậc",
        "No Formal quals < Lower Than A Level < A Level or Equivalent < HE Qualification < Post Graduate Qualification",
        "Ordinal",
        "Original",
    ),
    (
        "imd_band",
        "Demographic",
        "Ordinal",
        "Thứ bậc",
        "Unknown < 0-10% < 10-20 < 20-30% < ... < 90-100%",
        "Ordinal",
        "Original",
    ),
    (
        "age_band",
        "Demographic",
        "Ordinal",
        "Thứ bậc",
        "0-35 < 35-55 < 55<=",
        "Ordinal",
        "Original",
    ),
    (
        "disability",
        "Demographic",
        "Binary",
        "Nhị phân",
        "",
        "Binary 0/1 (Y=1,N=0)",
        "Original",
    ),
    (
        "num_of_prev_attempts",
        "Demographic",
        "Numeric (discrete)",
        "Định lượng (rời rạc)",
        "",
        "StandardScaler",
        "Original",
    ),
    (
        "studied_credits",
        "Demographic",
        "Numeric (discrete)",
        "Định lượng (rời rạc)",
        "",
        "StandardScaler",
        "Original",
    ),
    (
        "date_registration",
        "Demographic",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "StandardScaler",
        "Original",
    ),
    (
        "total_clicks",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "n_days_active",
        "Engagement",
        "Numeric (discrete)",
        "Định lượng (rời rạc)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "clicks_forumng",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "clicks_oucontent",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "clicks_resource",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "clicks_homepage",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "clicks_oucollaborate",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "clicks_quiz",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "clicks_subpage",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "clicks_url",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "max_clicks_single_day",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "mean_clicks_per_active_day",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "log1p + StandardScaler",
        "Derived",
    ),
    (
        "days_since_last_activity",
        "Engagement",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "winsorize + StandardScaler",
        "Derived",
    ),
    (
        "mean_score_to_date",
        "Performance",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "StandardScaler",
        "Derived",
    ),
    (
        "weighted_score_to_date",
        "Performance",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "winsorize + StandardScaler",
        "Derived",
    ),
    (
        "n_assessments_submitted",
        "Performance",
        "Numeric (discrete)",
        "Định lượng (rời rạc)",
        "",
        "StandardScaler",
        "Derived",
    ),
    (
        "not_submitted",
        "Performance",
        "Binary (indicator)",
        "Nhị phân (chỉ báo)",
        "",
        "Passthrough 0/1",
        "Derived",
    ),
    (
        "date_unregistration",
        "Temporal",
        "Numeric (continuous)",
        "Định lượng (liên tục)",
        "",
        "Not a feature (Withdrawn analysis)",
        "Original",
    ),
    (
        "module_presentation_length",
        "Temporal",
        "Numeric (discrete)",
        "Định lượng (rời rạc)",
        "",
        "Checkpoint-day conversion",
        "Original",
    ),
    (
        "final_result",
        "Target (raw)",
        "Nominal (raw)",
        "Danh định (gốc)",
        "",
        "Mapped to at_risk",
        "Original",
    ),
    (
        "at_risk",
        "Target",
        "Binary (target)",
        "Nhị phân (mục tiêu)",
        "",
        "None (target)",
        "Derived",
    ),
]

HEADERS = [
    "Variable",
    "Group / Nhóm",
    "Type (EN)",
    "Kiểu (VI)",
    "Order if ordinal / Thứ tự",
    "Encoding",
    "Origin / Nguồn gốc",
]

wb = Workbook()
ws = wb.active
ws.title = "VariableTyping"

header_fill = PatternFill("solid", fgColor="2166AC")
header_font = Font(bold=True, color="FFFFFF")
for j, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=j, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(vertical="center", wrap_text=True)

group_fill = {
    "Identifier": "EEEEEE",
    "Demographic": "E8F0FB",
    "Engagement": "FDEBE7",
    "Performance": "EAF3EA",
    "Temporal": "F3EDF7",
    "Target": "FFF6DD",
    "Target (raw)": "FFF6DD",
}
for i, row in enumerate(ROWS, start=2):
    for j, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    fill = group_fill.get(row[1], "FFFFFF")
    ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=fill)

widths = [26, 16, 22, 24, 46, 28, 16]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A2"

legend = wb.create_sheet("Legend")
legend_rows = [
    ["Column", "Meaning (EN)", "Ý nghĩa (VI)"],
    ["Type", "Statistical type of the variable", "Kiểu thống kê của biến"],
    [
        "Ordinal",
        "Has an intrinsic order (encode with OrdinalEncoder)",
        "Có thứ tự nội tại (mã hoá OrdinalEncoder)",
    ],
    [
        "Nominal",
        "No order (encode with OneHotEncoder)",
        "Không thứ tự (mã hoá OneHotEncoder)",
    ],
    ["Binary", "Two values mapped to 0/1", "Hai giá trị ánh xạ 0/1"],
    [
        "Origin",
        "Original column vs engineered feature",
        "Cột gốc hay đặc trưng phái sinh",
    ],
    [
        "Coverage",
        "30 variables, 100% of the master table",
        "30 biến, bao phủ 100% bảng hợp nhất",
    ],
]
for i, r in enumerate(legend_rows, 1):
    for j, v in enumerate(r, 1):
        cell = legend.cell(row=i, column=j, value=v)
        if i == 1:
            cell.fill = header_fill
            cell.font = header_font
for j, w in enumerate([14, 48, 48], 1):
    legend.column_dimensions[get_column_letter(j)].width = w

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print("wrote", OUT, "rows:", len(ROWS))
